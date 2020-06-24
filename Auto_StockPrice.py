import time
start_time = time.time()

from openpyxl import load_workbook
import time
#import bs4
import requests as rqs
from bs4 import BeautifulSoup
import pandas as pd
import re
import sys

def get_data(ticker, c, info):
    
    u = 'https://web.tmxmoney.com/quote.php?qm_symbol='
    

    if c == 'usd': # For US stocks
       url = u + ticker + ':us' 
    elif c == 'cad':
        url = u + ticker
    else:
        print("Make sure the currency is in 'CAD' or 'USD'")
        sys.exit(0)

    #print(url)
    response = rqs.get(url)
    soup = BeautifulSoup(response.text, 'lxml')
    
    # Get the Stock/ETF Price
    for pr in soup.find_all('span', {"class" : 'price'}):
        items = " ".join(pr.text.split()).replace(',','')
        p = re.findall("\d+\.\d+", items)
        price = float(p[0])

    # Get the Dividend Frequency
    for pr in soup.find_all('div', {'class':'dq-card'}):
        if pr(text=re.compile('Div. Frequency:')):
            freq = pr.strong.text
    #return freq
    
    # Get the Dividend
    for pr in soup.find_all('div', {'class':'dq-card'}):
        if pr(text=re.compile('Dividend')):
            div_ = pr.strong.text.split()[0]
            #div_ = float(div_)
            try:
                div_ = float(div_)
            except:
                print("No dividend was available for: {}.".format(ticker))
    
    try:
        if info == 'potential':
            return [price, freq, div_, url]  
        elif info == 'current':
            return [price, freq, div_, url] 
        else:
            print("Please pass 'potential' or 'current' to this function")
            sys.exit(0)
    except UnboundLocalError:
        print("Make sure the Ticker Symbol is correct: {} ".format(ticker))
    except:
        sys.exit(1)
    
def strip_whitespaces(d):
    d.columns = d.columns.str.strip()
    d['Ticker'] = d['Ticker'].str.strip()
    d['Currency'] = d['Currency'].str.strip()
    return d

def get_col_indx(d,clms):
    
    indexes = []
    
    for clm in clms:
        indexes.append(list(d.columns).index(clm) + 1)
        
    return indexes

if __name__ == "__main__":

    file_name = 'Stock_Analysis.xlsx'
    sheet_names = ['Potential_Investments', 'Current_Holdings']
    
    try:
        df = pd.read_excel(file_name, sheet_names[0], keep_default_na=False) # Store Potential_Investments sheet in a dataframe
        ch = pd.read_excel(file_name, sheet_names[1], keep_default_na=False) # Store Current_Holdings sheet in a dataframe
        
        wb = load_workbook(filename = file_name)
        
    except FileNotFoundError:
        print(f"No file named '{file_name}' in the directory")
        sys.exit(0)
        
    except:
        sys.exit(1)
    
    pot_inv_sheet = wb['Potential_Investments'] # Retrieve the Potential_Investments sheet
    curr_hold_sheet = wb['Current_Holdings']
    
    # Remove whitespaces from column titles and data values
    
    df = strip_whitespaces(df) # Potential Investments sheet
    df['Currency'] = df['Currency'].str.lower()
    
    ch = strip_whitespaces(ch)
    ch['Currency'] = ch['Currency'].str.lower()
    
    #sheets = (wb.sheetnames)
    
    pot_cols = ['Current_Price', 'Div_Frequency','Dividend','Link']
    curr_cols = ['Current_Price', 'Div_Frequency','Dividend','Link']
    
    pot_col_idxs = get_col_indx(df,pot_cols)
    curr_col_idxs = get_col_indx(ch, curr_cols)
    
    df[pot_cols] = df.apply(lambda x: get_data(x['Ticker'], x['Currency'], 'potential'), axis=1, result_type='expand')
    ch[curr_cols] = ch.apply(lambda x: get_data(x['Ticker'], x['Currency'], 'current'), axis=1, result_type='expand')
    
    #df['Stck_Names'] = df['Stock Name']
    df.set_index('Ticker', inplace=True) # Need to assign names as index for use later when adding in scraped data
    ch.set_index('Ticker', inplace=True)
    
    for rn in range(2,pot_inv_sheet.max_row+1):
        rowName = pot_inv_sheet.cell(row=rn, column = 1).value
        if rowName in set(df.index.values):
            pot_inv_sheet.cell(row=rn,column=pot_col_idxs[0]).value = df.at[rowName,'Current_Price'] # Requires stock names to be the row index
            pot_inv_sheet.cell(row=rn,column=pot_col_idxs[1]).value = df.at[rowName,'Div_Frequency']
            pot_inv_sheet.cell(row=rn,column=pot_col_idxs[2]).value = df.at[rowName,'Dividend']
            pot_inv_sheet.cell(row=rn,column=pot_col_idxs[3]).value = df.at[rowName,'Link']
            
    for rn in range(2,curr_hold_sheet.max_row+1):
        rowName = curr_hold_sheet.cell(row=rn, column = 1).value
        if rowName in set(ch.index.values):
            curr_hold_sheet.cell(row=rn,column=curr_col_idxs[0]).value = ch.at[rowName,'Current_Price'] # Requires stock names to be the row index
            curr_hold_sheet.cell(row=rn,column=curr_col_idxs[1]).value = ch.at[rowName,'Div_Frequency']
            curr_hold_sheet.cell(row=rn,column=curr_col_idxs[2]).value = ch.at[rowName,'Dividend']
            curr_hold_sheet.cell(row=rn,column=curr_col_idxs[3]).value = ch.at[rowName,'Link']

    
    wb.save(file_name)
    
    print("--- %s seconds ---" % (time.time() - start_time))
    #test



